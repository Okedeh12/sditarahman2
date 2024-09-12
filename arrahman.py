from openpyxl import Workbook
from openpyxl.drawing.image import Image
import streamlit as st
import pandas as pd
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from fpdf import FPDF
from streamlit_option_menu import option_menu
from datetime import datetime

# Define file paths
CSV_PEMBAYARAN_SPP = 'data/pembayaran_spp.csv'
CSV_GAJI_GURU = 'data/gaji_guru.csv'
CSV_DAFTAR_ULANG = 'data/daftar_ulang.csv'
CSV_PENGELUARAN = 'data/pengeluaran.csv'
PERSISTENT_DIR = 'data/uploads'
LOGO_PATH = 'data/HN.png'

# Ensure data directories exist
os.makedirs('data', exist_ok=True)
os.makedirs(PERSISTENT_DIR, exist_ok=True)

def save_pembayaran_spp(nama_siswa, kelas, bulan, jumlah, biaya_spp):
    df = pd.DataFrame([[nama_siswa, kelas, bulan, jumlah, biaya_spp]], columns=['nama_siswa', 'kelas', 'bulan', 'jumlah', 'biaya_spp'])
    if os.path.exists(CSV_PEMBAYARAN_SPP):
        df_existing = pd.read_csv(CSV_PEMBAYARAN_SPP)
        df = pd.concat([df_existing, df], ignore_index=True)
    df.to_csv(CSV_PEMBAYARAN_SPP, index=False)

def save_gaji_guru(nama_guru, bulan_gaji, gaji, tunjangan):
    df = pd.DataFrame([[nama_guru, bulan_gaji, gaji, tunjangan]], columns=['nama_guru', 'bulan_gaji', 'gaji', 'tunjangan'])
    if os.path.exists(CSV_GAJI_GURU):
        df_existing = pd.read_csv(CSV_GAJI_GURU)
        df = pd.concat([df_existing, df], ignore_index=True)
    df.to_csv(CSV_GAJI_GURU, index=False)

def save_daftar_ulang(nama_siswa, kelas, biaya_daftar_ulang, pembayaran, tahun):
    df = pd.DataFrame([[nama_siswa, kelas, biaya_daftar_ulang, pembayaran, tahun]], columns=['nama_siswa', 'kelas', 'biaya_daftar_ulang', 'pembayaran', 'tahun'])
    if os.path.exists(CSV_DAFTAR_ULANG):
        df_existing = pd.read_csv(CSV_DAFTAR_ULANG)
        df = pd.concat([df_existing, df], ignore_index=True)
    df.to_csv(CSV_DAFTAR_ULANG, index=False)

def save_pengeluaran(nama_penerima, keterangan_biaya, total_biaya, file_path=None):
    df = pd.DataFrame([[nama_penerima, keterangan_biaya, total_biaya, file_path]], columns=['nama_penerima', 'keterangan_biaya', 'total_biaya', 'file_path'])
    if os.path.exists(CSV_PENGELUARAN):
        df_existing = pd.read_csv(CSV_PENGELUARAN)
        df = pd.concat([df_existing, df], ignore_index=True)
    df.to_csv(CSV_PENGELUARAN, index=False)

def generate_receipt(nama_siswa, kelas, bulan, jumlah, biaya_spp, receipt_type):
    pdf = FPDF()
    pdf.add_page()
    
    # School details
    school_name = "SD IT ARAHMAN"
    school_address = "JATIMULYO"

    # Add logo
    if os.path.exists(LOGO_PATH):
        pdf.image(LOGO_PATH, x=10, y=8, w=33)
    else:
        pdf.set_font("Arial", size=12)
        pdf.cell(0, 10, txt="Logo Tidak Ditemukan", ln=True, align='C')

    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, txt=school_name, ln=True, align='C')
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, txt=school_address, ln=True, align='C')
    pdf.ln(10)
    
    if receipt_type == "spp":
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, txt="Kwitansi Pembayaran SPP", ln=True, align='C')
        pdf.ln(10)
        details = [
            ("Nama Siswa", nama_siswa),
            ("Kelas", kelas),
            ("Bulan", bulan),
            ("Jumlah Pembayaran", f"Rp {jumlah:,}"),
            ("Biaya SPP per Bulan", f"Rp {biaya_spp:,}"),
            ("Tanggal", datetime.now().strftime('%Y-%m-%d'))
        ]
    elif receipt_type == "daftar_ulang":
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, txt="Kwitansi Pembayaran Daftar Ulang", ln=True, align='C')
        pdf.ln(10)
        details = [
            ("Nama Siswa", nama_siswa),
            ("Kelas", kelas),
            ("Biaya Daftar Ulang", f"Rp {biaya_spp:,}"),
            ("Pembayaran", f"Rp {jumlah:,}"),
            ("Tanggal", datetime.now().strftime('%Y-%m-%d'))
        ]
    elif receipt_type == "gaji":
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, txt="Kwitansi Pembayaran Gaji Guru", ln=True, align='C')
        pdf.ln(10)
        details = [
            ("Nama Guru", nama_siswa),
            ("Bulan Gaji", kelas),
            ("Gaji", f"Rp {jumlah:,}"),
            ("Tunjangan", f"Rp {biaya_spp:,}"),
            ("Tanggal", datetime.now().strftime('%Y-%m-%d'))
        ]
    else:
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, txt="Kwitansi Pengeluaran", ln=True, align='C')
        pdf.ln(10)
        details = [
            ("Nama Penerima", nama_siswa),
            ("Keterangan Biaya", kelas),
            ("Total Biaya", f"Rp {jumlah:,}"),
            ("Tanggal", datetime.now().strftime('%Y-%m-%d'))
        ]
    
    pdf.set_font("Arial", size=12)
    for label, value in details:
        pdf.cell(100, 10, txt=label, border=1)
        pdf.cell(90, 10, txt=f": {value}", border=1, ln=True)
    
    pdf.ln(10)
    pdf.cell(0, 10, txt="Tanda Terima", ln=True, align='R')

    pdf_output = BytesIO()
    pdf_output.write(pdf.output(dest='S').encode('latin1'))
    pdf_output.seek(0)

    return pdf_output

def generate_expense_receipt(nama_penerima, keterangan_biaya, total_biaya):
    return generate_receipt(nama_penerima, keterangan_biaya, '', total_biaya, total_biaya, 'pengeluaran')

def load_data():
    df_spp = pd.read_csv(CSV_PEMBAYARAN_SPP) if os.path.exists(CSV_PEMBAYARAN_SPP) else pd.DataFrame()
    df_gaji = pd.read_csv(CSV_GAJI_GURU) if os.path.exists(CSV_GAJI_GURU) else pd.DataFrame()
    df_daftar_ulang = pd.read_csv(CSV_DAFTAR_ULANG) if os.path.exists(CSV_DAFTAR_ULANG) else pd.DataFrame()
    df_pengeluaran = pd.read_csv(CSV_PENGELUARAN) if os.path.exists(CSV_PENGELUARAN) else pd.DataFrame()
    return df_spp, df_gaji, df_daftar_ulang, df_pengeluaran

def handle_csv_upload(uploaded_csv_file):
    try:
        new_data = pd.read_csv(uploaded_csv_file)
        new_data.to_csv(CSV_PENGELUARAN, mode='a', header=False, index=False)
        df_pengeluaran = pd.read_csv(CSV_PENGELUARAN)
        st.success("File CSV berhasil diupload dan ditambahkan ke data pengeluaran!")
        st.dataframe(df_pengeluaran)
    except Exception as e:
        st.error(f"Terjadi kesalahan: {e}")

def handle_photo_upload(uploaded_file, row_index):
    if uploaded_file:
        file_path = os.path.join(PERSISTENT_DIR, uploaded_file.name)
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getvalue())
        df = pd.read_csv(CSV_PENGELUARAN)
        df.at[row_index, 'file_path'] = file_path
        df.to_csv(CSV_PENGELUARAN, index=False)
        st.success("Foto berhasil diupload!")

def export_to_excel_with_photos(df_spp, df_gaji, df_daftar_ulang, df_pengeluaran):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from io import BytesIO
    
    output = BytesIO()
    workbook = Workbook()
    
    # Helper function to add dataframe to workbook sheet
    def add_df_to_sheet(sheet, df, sheet_name):
        for r_idx, row in enumerate(df.itertuples(), 1):
            for c_idx, value in enumerate(row[1:], 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        sheet.title = sheet_name

    # Create sheets for each dataframe
    add_df_to_sheet(workbook.create_sheet(), df_spp, 'Pembayaran SPP')
    add_df_to_sheet(workbook.create_sheet(), df_gaji, 'Gaji Guru')
    add_df_to_sheet(workbook.create_sheet(), df_daftar_ulang, 'Daftar Ulang')
    
    # Add Pengeluaran data with photos
    pengeluaran_sheet = workbook.create_sheet(title='Pengeluaran')
    for r_idx, row in enumerate(df_pengeluaran.itertuples(), 1):
        for c_idx, value in enumerate(row[1:], 1):
            pengeluaran_sheet.cell(row=r_idx, column=c_idx, value=value)
        if row.file_path and os.path.exists(row.file_path):
            img = Image(row.file_path)
            img.width = 200  # Resize as needed
            img.height = 150
            pengeluaran_sheet.add_image(img, f'E{r_idx}')

    workbook.save(output)
    output.seek(0)
    
    return output.getvalue()

st.title("Aplikasi Manajemen Keuangan Sekolah")

# Sidebar Menu
with st.sidebar:
    selected = option_menu(
        menu_title="Menu",
        options=["Upload CSV", "Pembayaran SPP", "Gaji Guru", "Daftar Ulang", "Pengeluaran", "Laporan Keuangan"],
        icons=["upload", "cash", "person", "list", "file-earmark-text", "file-earmark-excel"],
        menu_icon="cast",
        default_index=0,
    )

df_spp, df_gaji, df_daftar_ulang, df_pengeluaran = load_data()

if selected == "Upload CSV":
    st.subheader("Upload File CSV Pengeluaran")
    uploaded_csv_file = st.file_uploader("Pilih file CSV", type=["csv"], key="upload_pengeluaran_csv")
    if uploaded_csv_file:
        handle_csv_upload(uploaded_csv_file)

elif selected == "Pembayaran SPP":
    st.subheader("Data Pembayaran SPP")
    if not df_spp.empty:
        selected_rows = st.multiselect(
            "Pilih data untuk diunduh kwitansi:",
            df_spp.index,
            format_func=lambda x: f"{df_spp.iloc[x]['nama_siswa']} ({df_spp.iloc[x]['bulan']})"
        )
        for index in selected_rows:
            row = df_spp.iloc[index]
            receipt = generate_receipt(row.get('nama_siswa', ''), row.get('kelas', ''), row.get('bulan', ''), row.get('jumlah', 0), row.get('biaya_spp', 0), 'spp')
            st.download_button(
                label=f"Download Kwitansi {row.get('nama_siswa', '')} ({row.get('bulan', '')})",
                data=receipt,
                file_name=f"kwitansi_spp_{row.get('nama_siswa', '')}_{row.get('bulan', '')}.pdf",
                mime="application/pdf",
                key=f"download_spp_{index}"
            )
    else:
        st.info("Tidak ada data Pembayaran SPP.")

elif selected == "Gaji Guru":
    st.subheader("Data Gaji Guru")
    if not df_gaji.empty:
        selected_rows = st.multiselect(
            "Pilih data untuk diunduh kwitansi:",
            df_gaji.index,
            format_func=lambda x: f"{df_gaji.iloc[x]['nama_guru']} ({df_gaji.iloc[x]['bulan_gaji']})"
        )
        for index in selected_rows:
            row = df_gaji.iloc[index]
            receipt = generate_receipt(row.get('nama_guru', ''), row.get('bulan_gaji', ''), '', row.get('gaji', 0), row.get('tunjangan', 0), 'gaji')
            st.download_button(
                label=f"Download Kwitansi {row.get('nama_guru', '')} ({row.get('bulan_gaji', '')})",
                data=receipt,
                file_name=f"kwitansi_gaji_{row.get('nama_guru', '')}_{row.get('bulan_gaji', '')}.pdf",
                mime="application/pdf",
                key=f"download_gaji_{index}"
            )
    else:
        st.info("Tidak ada data Gaji Guru.")

elif selected == "Daftar Ulang":
    st.subheader("Data Daftar Ulang")
    if not df_daftar_ulang.empty:
        selected_rows = st.multiselect(
            "Pilih data untuk diunduh kwitansi:",
            df_daftar_ulang.index,
            format_func=lambda x: f"{df_daftar_ulang.iloc[x]['nama_siswa']} ({df_daftar_ulang.iloc[x]['tahun']})"
        )
        for index in selected_rows:
            row = df_daftar_ulang.iloc[index]
            receipt = generate_receipt(row.get('nama_siswa', ''), row.get('kelas', ''), '', row.get('pembayaran', 0), row.get('biaya_daftar_ulang', 0), 'daftar_ulang')
            st.download_button(
                label=f"Download Kwitansi {row.get('nama_siswa', '')} ({row.get('tahun', '')})",
                data=receipt,
                file_name=f"kwitansi_daftar_ulang_{row.get('nama_siswa', '')}_{row.get('tahun', '')}.pdf",
                mime="application/pdf",
                key=f"download_daftar_ulang_{index}"
            )
    else:
        st.info("Tidak ada data Daftar Ulang.")

elif selected == "Pengeluaran":
    st.title("Pengelolaan Pengeluaran")
    with st.form("pengeluaran_form"):
        nama_penerima = st.text_input("Nama Penerima", key="pengeluaran_nama_penerima")
        keterangan_biaya = st.text_input("Keterangan Biaya", key="pengeluaran_keterangan_biaya")
        total_biaya = st.number_input("Total Biaya", min_value=0, key="pengeluaran_total_biaya")
        uploaded_file = st.file_uploader("Upload Foto Bukti Pengeluaran (opsional)", type=["jpg", "jpeg", "png"], key="upload_pengeluaran_image")
        submitted = st.form_submit_button("Simpan")

        if submitted:
            file_path = None
            if uploaded_file:
                file_path = os.path.join(PERSISTENT_DIR, uploaded_file.name)
                with open(file_path, "wb") as f:
                    f.write(uploaded_file.getvalue())
            
            save_pengeluaran(nama_penerima, keterangan_biaya, total_biaya, file_path)
            df_pengeluaran = pd.read_csv(CSV_PENGELUARAN)
            st.success("Pengeluaran berhasil disimpan!")

    st.write("**Data Pengeluaran**")
    st.dataframe(df_pengeluaran)

    st.write("**Download Kwitansi Pengeluaran**")
    if not df_pengeluaran.empty:
        for index, row in df_pengeluaran.iterrows():
            receipt = generate_expense_receipt(row.get('nama_penerima', ''), row.get('keterangan_biaya', ''), row.get('total_biaya', 0))
            st.download_button(
                label=f"Download Kwitansi {row.get('nama_penerima', '')}",
                data=receipt,
                file_name=f"kwitansi_pengeluaran_{row.get('nama_penerima', '')}.pdf",
                mime="application/pdf",
                key=f"download_pengeluaran_{index}"
            )
            if row.file_path and os.path.exists(row.file_path):
                with open(row.file_path, "rb") as photo_file:
                    st.download_button(
                        label=f"Download Foto {row.get('nama_penerima', '')}",
                        data=photo_file,
                        file_name=os.path.basename(row.file_path),
                        mime="image/jpeg",
                        key=f"download_foto_{index}"
                    )

elif selected == "Laporan Keuangan":
    st.title("Laporan Keuangan")
    
    # Display dataframes
    st.write("**Laporan Pembayaran SPP**")
    st.dataframe(df_spp)

    st.write("**Laporan Gaji Guru**")
    st.dataframe(df_gaji)

    st.write("**Laporan Daftar Ulang**")
    st.dataframe(df_daftar_ulang)

    st.write("**Laporan Pengeluaran**")
    st.dataframe(df_pengeluaran)

    # Export to Excel with Photos
    excel_data = export_to_excel_with_photos(df_spp, df_gaji, df_daftar_ulang, df_pengeluaran)
    st.download_button(
        label="Download Excel File with Photos",
        data=excel_data,
        file_name="laporan_keuangan_dengan_foto.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_excel_with_photos"
    )
